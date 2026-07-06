"""
Report generation — PDF (ReportLab) and Excel (openpyxl) exports of threat logs.
"""
from __future__ import annotations

import io
import time
from datetime import datetime
from typing import List

from app.db.models.event import ThreatLog


def _fmt_ts(ts: float) -> str:
    return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")


def build_excel_report(threats: List[ThreatLog]) -> bytes:
    """Return an .xlsx workbook of the given threats as bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Threats"
    headers = ["ID", "Timestamp", "Camera", "Category", "Severity",
               "Score", "Message", "Acknowledged"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for t in threats:
        ws.append([
            t.id, _fmt_ts(t.timestamp), t.camera_id, t.category,
            t.severity, round(t.score, 3), t.message, t.acknowledged,
        ])
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_report(threats: List[ThreatLog], title: str = "SentinelAI Threat Report"
                     ) -> bytes:
    """Return a PDF report of the given threats as bytes."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=title)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles["Title"]),
        Paragraph(f"Generated: {_fmt_ts(time.time())}", styles["Normal"]),
        Paragraph(f"Total events: {len(threats)}", styles["Normal"]),
        Spacer(1, 16),
    ]

    data = [["Time", "Camera", "Category", "Severity", "Score", "Message"]]
    for t in threats[:500]:  # keep PDFs bounded
        data.append([
            _fmt_ts(t.timestamp), t.camera_id, t.category,
            t.severity, f"{t.score:.2f}", t.message[:48],
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F3F4F6")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()
